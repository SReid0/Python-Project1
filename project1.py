import openpyxl

# Open Excel file (named graduates)
workbook = openpyxl.load_workbook('graduates.xlsx')

# Select correct sheet
worksheet = workbook['Sheet1']

# Get user input
state = input("Enter a state name ('all' to view all states, 'top10' to view top 10 states): ")

# Check if user input is all
if state.lower() == "all":
    # Loop through and print all rows if input is all
    states_list = []
    for row in worksheet.iter_rows(min_row=2, max_col=3):
        state_name = row[0].value
        # Graduates per year (row_b) and total enrollment (row_c) for given state
        row_b = row[1].value
        row_c = row[2].value
        # Check if either row is empty to avoid errors
        if row_b is None or row_c is None:
            percent_grad = 0
        elif row_c == 0:
            percent_grad = 0
        else:
            percent_grad = row_b / row_c * 100
        # Check if state is empty to avoid errors
        if state_name is not None:
            states_list.append((state_name, percent_grad))
            print(f"State: {state_name}")
            print(f"Enrollment for {state_name}: {row_c}")
            print(f"Graduates per year for {state_name}: {row_b}")
            print(f"Percentage of Graduates for {state_name}: {percent_grad:.2f}%")

    print(f"These percentages are based off 2022 Graduates per year for each state and Spring 2022 Enrollment of each state.")

    # Check if user input is top10
elif state.lower() == "top10":
    # Create a list of all states for later calculation
    states_list = []
    for row in worksheet.iter_rows(min_row=2, max_col=3):
        state_name = row[0].value
        # Graduates per year (row_b) and total enrollment (row_c) for given state
        row_b = row[1].value
        row_c = row[2].value
        # Check if either row is empty to avoid errors
        if row_b is None or row_c is None:
            percent_grad = 0
        elif row_c == 0:
            percent_grad = 0
        else:
            percent_grad = row_b / row_c * 100
        # Check if state is empty to avoid errors
        if state_name is not None:
            states_list.append((state_name, percent_grad))
    # Sort states list in descending order based on percentage
    states_list.sort(key=lambda x: x[1], reverse=True)
    # Print top 10 states based on percentage
    for i in range(min(10, len(states_list))):
        state_name, percent_grad = states_list[i]
        print(f"{i+1}. State: {state_name} - Percentage of Graduates: {percent_grad:.2f}%")

    print(f"These percentages are based off 2022 Graduates per year for each state and Spring 2022 Enrollment of each state.")

else:
    # Find the row of the state based off user input
    state_row = None
    for row in worksheet.iter_rows(min_row=2, max_col=1):
        if row[0].value.lower() == state.lower():
            state_row = row[0].row
            break

    # If state not found print error and exit
    if state_row is None:
        print(f"Error: State {state} not found")
        exit()

    # Get graduates per year (row_b) and total enrollment (row_c) for given state
    row_b = worksheet.cell(row=state_row, column=2).value
    row_c = worksheet.cell(row=state_row, column=3).value

    # Calculate percentage
    # Check if either row is empty before calculation to avoid errors
    # Graduates per year (row_b) and total enrollment (row_c) for given state
    if row_b is None or row_c is None:
        percent_grad = 0
    elif row_c == 0:
        percent_grad = 0
    else:
        percent_grad = row_b / row_c * 100

    # The states are separated by a space so use split to capatilize both words ie... New York
    state_parts = state.split(' ')
    state_parts = [part.capitalize() for part in state_parts]
    state_capitalized = ' '.join(state_parts)

    # Print results
    print(f"State: {state_capitalized}")
    print(f"Percentage of Graduates for {state_capitalized}: {percent_grad:.2f}%")
    print(f"These percentages are based off 2022 Graduates per year for each state and Spring 2022 Enrollment of each state.")